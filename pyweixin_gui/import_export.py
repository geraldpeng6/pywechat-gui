from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .models import FileBatchRow, MessageBatchRow, RelayRouteRow, TaskType

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - optional dependency in non-GUI envs
    Workbook = None
    load_workbook = None


MESSAGE_HEADERS = [
    "enabled",
    "session_name",
    "message",
    "at_members",
    "at_all",
    "clear_before_send",
    "send_delay_sec",
    "remark",
]

FILE_HEADERS = [
    "enabled",
    "session_name",
    "file_paths",
    "with_message",
    "message",
    "message_first",
    "remark",
]

ROUTE_HEADERS = [
    "downstream_session",
]

ROUTE_TEMPLATE_HEADERS = [
    "收件人会话",
]

ROUTE_HEADER_ALIASES = {
    "enabled": "enabled",
    "启用": "enabled",
    "upstream_session": "upstream_session",
    "上游会话": "upstream_session",
    "source_session": "upstream_session",
    "downstream_session": "downstream_session",
    "下游会话": "downstream_session",
    "收件人会话": "downstream_session",
    "收件人": "downstream_session",
    "目标会话": "downstream_session",
    "target_session": "downstream_session",
    "downstream_sessions": "downstream_sessions",
    "下游会话列表": "downstream_sessions",
    "收件人列表": "downstream_sessions",
    "收件人会话列表": "downstream_sessions",
    "remark": "remark",
    "备注": "remark",
}


def headers_for(task_type: TaskType) -> list[str]:
    return MESSAGE_HEADERS if task_type is TaskType.MESSAGE else FILE_HEADERS


def _decode_text_file(path: Path, encodings: tuple[str, ...] = ("utf-8-sig", "gbk")) -> str:
    raw = path.read_bytes()
    for encoding in encodings:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("text", raw, 0, 1, "无法识别文件编码")


def _load_csv_rows(path: Path) -> list[dict[str, Any]]:
    text = _decode_text_file(path)
    reader = csv.DictReader(text.splitlines())
    return [dict(row) for row in reader]


def _load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("未安装 openpyxl，无法导入 .xlsx 文件")
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        items.append({header[index]: row[index] for index in range(len(header))})
    return items


def load_rows(task_type: TaskType, path: str | Path) -> list[MessageBatchRow] | list[FileBatchRow]:
    rows = load_table_rows(path)
    if task_type is TaskType.MESSAGE:
        return [MessageBatchRow.from_mapping(row) for row in rows]
    return [FileBatchRow.from_mapping(row) for row in rows]


def load_table_rows(path: str | Path) -> list[dict[str, Any]]:
    file_path = Path(path)
    if file_path.suffix.lower() == ".csv":
        return _load_csv_rows(file_path)
    if file_path.suffix.lower() == ".xlsx":
        return _load_xlsx_rows(file_path)
    raise ValueError("仅支持导入 .csv 或 .xlsx 文件")


def load_route_rows(path: str | Path) -> list[RelayRouteRow]:
    rows = load_table_rows(path)
    normalized = [_normalize_route_mapping(row) for row in rows]
    expanded: list[RelayRouteRow] = []
    for row in normalized:
        expanded.extend(_expand_route_mapping(row))
    return expanded


def dump_table(headers: list[str], rows: list[dict[str, Any]], path: str | Path) -> None:
    file_path = Path(path)
    if file_path.suffix.lower() == ".csv":
        with file_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({key: row.get(key, "") for key in headers})
        return
    if file_path.suffix.lower() == ".xlsx":
        if Workbook is None:
            raise RuntimeError("未安装 openpyxl，无法导出 .xlsx 文件")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(key, "") for key in headers])
        workbook.save(file_path)
        return
    raise ValueError("仅支持导出 .csv 或 .xlsx 文件")


def dump_rows(task_type: TaskType, rows: list[dict[str, Any]], path: str | Path) -> None:
    dump_table(headers_for(task_type), rows, path)


def dump_route_rows(rows: list[dict[str, Any]], path: str | Path) -> None:
    template_rows = []
    for row in rows:
        template_rows.append(
            {
                "收件人会话": row.get("downstream_session", ""),
            }
        )
    dump_table(ROUTE_TEMPLATE_HEADERS, template_rows, path)


def load_session_names(path: str | Path) -> list[str]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".txt":
        text = _decode_text_file(file_path)
        return [line.strip() for line in text.splitlines() if line.strip()]
    if suffix == ".csv":
        rows = _load_csv_rows(file_path)
        return _extract_session_names(rows)
    if suffix == ".xlsx":
        rows = _load_xlsx_rows(file_path)
        return _extract_session_names(rows)
    raise ValueError("会话名单仅支持导入 .txt、.csv 或 .xlsx 文件")


def _extract_session_names(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return []
    first_row = rows[0]
    if "session_name" in first_row:
        return [str(row.get("session_name", "")).strip() for row in rows if str(row.get("session_name", "")).strip()]
    values: list[str] = []
    for row in rows:
        for value in row.values():
            text = str(value or "").strip()
            if text:
                values.append(text)
                break
    return values


def _normalize_route_mapping(mapping: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in mapping.items():
        text_key = str(key or "").strip()
        canonical = ROUTE_HEADER_ALIASES.get(text_key)
        if canonical:
            normalized[canonical] = value
    return normalized


def _expand_route_mapping(mapping: dict[str, Any]) -> list[RelayRouteRow]:
    downstream = str(mapping.get("downstream_session", "") or "").strip()
    downstreams = str(mapping.get("downstream_sessions", "") or "").strip()
    if downstreams:
        targets = [item.strip() for item in downstreams.replace("\n", "|").replace("，", "|").replace(",", "|").split("|") if item.strip()]
        return [
            RelayRouteRow.from_mapping(
                {
                    "downstream_session": target,
                }
            )
            for target in targets
        ]
    return [RelayRouteRow.from_mapping(mapping | {"downstream_session": downstream})]
