from __future__ import annotations

import importlib.util
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from pyweixin_gui.export_service import ChatExportService
from pyweixin_gui.models import ChatExportRequest, RuntimeOptions


class FakeAdapter:
    def dump_chat_history(self, session_name, number, options):
        return ["第一条消息", "第二条消息"], ["昨天 12:00", "昨天 12:01"]

    def save_chat_files(self, session_name, number, target_folder, options):
        folder = Path(target_folder)
        folder.mkdir(parents=True, exist_ok=True)
        (folder / "报告.pdf").write_text("demo", encoding="utf-8")
        return [str(folder / "报告.pdf")]

    def save_chat_media(self, session_name, number, target_folder, options):
        folder = Path(target_folder)
        folder.mkdir(parents=True, exist_ok=True)
        (folder / "截图.png").write_text("img", encoding="utf-8")
        (folder / "演示视频.mp4").write_text("video", encoding="utf-8")
        return [str(folder / "截图.png"), str(folder / "演示视频.mp4")]


class ParseFailureAdapter(FakeAdapter):
    def save_chat_files(self, session_name, number, target_folder, options):
        raise AttributeError("'NoneType' object has no attribute 'group'")


class ExportServiceTestCase(unittest.TestCase):
    def test_export_bundle_uses_unique_folder_when_same_second_repeats(self):
        service = ChatExportService(FakeAdapter())

        class FixedDateTime:
            @staticmethod
            def now():
                return datetime(2026, 4, 3, 9, 30, 0)

        with tempfile.TemporaryDirectory() as tempdir, patch("pyweixin_gui.paths.datetime", FixedDateTime):
            request = ChatExportRequest(
                session_name="测试群",
                target_folder=tempdir,
                export_messages=False,
                export_files=True,
                export_images=False,
                file_limit=10,
            )
            first = service.export_chat_bundle(request, RuntimeOptions())
            second = service.export_chat_bundle(request, RuntimeOptions())

        self.assertNotEqual(first.export_folder, second.export_folder)
        self.assertTrue(Path(first.export_folder).name.startswith("测试群-20260403-093000"))
        self.assertTrue(Path(second.export_folder).name.startswith("测试群-20260403-093000"))

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_export_bundle_writes_summary_and_messages(self):
        service = ChatExportService(FakeAdapter())
        with tempfile.TemporaryDirectory() as tempdir:
            request = ChatExportRequest(
                session_name="测试群",
                target_folder=tempdir,
                export_messages=True,
                export_files=True,
                export_images=False,
                message_limit=10,
                file_limit=10,
            )
            result = service.export_chat_bundle(request, RuntimeOptions())
            export_folder = Path(result.export_folder)
            self.assertTrue(export_folder.exists())
            self.assertTrue(Path(result.messages_xlsx).exists())
            self.assertEqual(result.message_count, 2)
            self.assertEqual(result.file_count, 1)
            self.assertEqual(result.media_count, 0)
            self.assertEqual(result.warnings, [])
            self.assertEqual(Path(result.messages_xlsx).name, "聊天记录.xlsx")
            self.assertEqual(Path(result.files_folder).name, "聊天文件")

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_export_bundle_keeps_messages_when_chat_file_parse_fails(self):
        service = ChatExportService(ParseFailureAdapter())
        with tempfile.TemporaryDirectory() as tempdir:
            request = ChatExportRequest(
                session_name="好友A",
                target_folder=tempdir,
                export_messages=True,
                export_files=True,
                export_images=False,
                message_limit=10,
                file_limit=10,
            )
            result = service.export_chat_bundle(request, RuntimeOptions())
            self.assertEqual(result.message_count, 2)
            self.assertEqual(result.file_count, 0)
            self.assertEqual(result.media_count, 0)
            self.assertTrue(result.messages_xlsx)
            self.assertTrue(result.files_folder)
            self.assertTrue(any("聊天文件未整理到结果中" in warning for warning in result.warnings))

    def test_export_bundle_includes_media_when_requested(self):
        service = ChatExportService(FakeAdapter())
        with tempfile.TemporaryDirectory() as tempdir:
            request = ChatExportRequest(
                session_name="好友A",
                target_folder=tempdir,
                export_messages=False,
                export_files=False,
                export_images=True,
                file_limit=10,
            )
            result = service.export_chat_bundle(request, RuntimeOptions())

        self.assertEqual(result.message_count, 0)
        self.assertEqual(result.file_count, 0)
        self.assertEqual(result.media_count, 2)
        self.assertTrue(result.media_folder)
        self.assertEqual(Path(result.media_folder).name, "聊天图片与视频")

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_export_bundle_orders_messages_from_oldest_to_newest(self):
        class ReverseAdapter(FakeAdapter):
            def dump_chat_history(self, session_name, number, options):
                return ["较新的消息", "较早的消息"], ["今天 10:01", "今天 10:00"]

        service = ChatExportService(ReverseAdapter())
        with tempfile.TemporaryDirectory() as tempdir:
            result = service.export_chat_bundle(
                ChatExportRequest(
                    session_name="好友A",
                    target_folder=tempdir,
                    export_messages=True,
                    export_files=False,
                    message_limit=10,
                    file_limit=10,
                ),
                RuntimeOptions(),
            )
            from openpyxl import load_workbook

            workbook = load_workbook(result.messages_xlsx)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            self.assertEqual(rows[1][2], "较早的消息")
            self.assertEqual(rows[2][2], "较新的消息")


if __name__ == "__main__":
    unittest.main()
